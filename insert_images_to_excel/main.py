import pathlib
import tkinter
import tkinter.filedialog

import openpyxl

import gui
import settings


def create_filenames_and_filepaths(dir: pathlib.Path) -> dict:

    files_in_dir = {
        filename.stem: filename for filename in dir.iterdir(
        ) if filename.is_file()
        }

    return files_in_dir


def insert_resized_image_to_cell(
    worksheet: openpyxl.worksheet,
    image: pathlib.Path,
    cell_address: str,
    width=100,
    height=100
        ) -> None:

    image = openpyxl.drawing.image.Image(image)

    worksheet.add_image(image, anchor=cell_address)

    image.width = width
    image.height = height


def set_cell_range_dimensions(
    worksheet: openpyxl.worksheet,
    cell_rows: int or range,
    cell_column: str,
    width: int or float,
    height: int or float
        ) -> None:

    for row in cell_rows:
        worksheet.row_dimensions[row].height = height
        worksheet.column_dimensions[cell_column].width = width


def insert_image_to_matched_cell_value(
    worksheet: openpyxl.worksheet,
    images_in_dir: dict,
    target_column: str,
    lookup_column: str,
    start_row: int
        ) -> None or dict:

    cells_with_images_not_found = {}

    for row in range(start_row, worksheet.max_row + 1):

        target_cell_address = target_column + str(row)
        lookup_column_index = openpyxl.utils.column_index_from_string(
            lookup_column
            )
        image = images_in_dir.get(
            worksheet.cell(row, lookup_column_index).value
            )

        if image:
            insert_resized_image_to_cell(
                worksheet,
                image=image,
                cell_address=target_cell_address)
        else:
            cells_with_images_not_found.update(
                {worksheet.cell(
                    row, lookup_column_index
                    ).value: target_cell_address})

    if cells_with_images_not_found:
        return cells_with_images_not_found


def main():

    tkinter.Tk().withdraw()

    PATH_TO_IMAGES_DIRECTORY = pathlib.Path(
        tkinter.filedialog.askdirectory(
            title='Choose directory with images',
            initialdir=pathlib.os.getcwd()))

    PATH_TO_TARGET_FILE = pathlib.Path(
        tkinter.filedialog.askopenfilename(
            title='Choose target file',
            initialdir=pathlib.os.getcwd()))

    PATH_TO_UPDATED_FILE = pathlib.Path(
        PATH_TO_TARGET_FILE.with_name(
            PATH_TO_TARGET_FILE.stem + '_UPD.xlsx'))

    LOOKUP_COLUMN = gui.create_simple_dialog(
        title='Setup',
        prompt='Enter column letter for lookup: ').upper()
    START_ROW = int(gui.create_simple_dialog(
        title='Setup',
        prompt='Enter start row: '))
    TARGET_COLUMN = 'A'

    workbook = openpyxl.load_workbook(PATH_TO_TARGET_FILE)
    worksheet = workbook.active

    set_cell_range_dimensions(
        worksheet,
        cell_rows=range(START_ROW, worksheet.max_row + 1),
        cell_column=TARGET_COLUMN,
        width=settings.CELL_WIDTH,
        height=settings.CELL_HEIGHT)

    insert_image_to_matched_cell_value(
        worksheet,
        images_in_dir=create_filenames_and_filepaths(
            PATH_TO_IMAGES_DIRECTORY),
        target_column=TARGET_COLUMN,
        lookup_column=LOOKUP_COLUMN,
        start_row=START_ROW
        )
    workbook.save(PATH_TO_UPDATED_FILE)
    workbook.close()
